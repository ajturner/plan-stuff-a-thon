#!/usr/bin/env python3
"""
create_survey.py
Generate the Troop 380 "family camping ideas" Survey123 form and publish it
to ArcGIS Online. Survey123 will auto-create a hosted feature layer for the
responses; that layer is what the web app reads from once you swap data.js
to a remote source.

Workflow it sets up
-------------------
    1. Family fills out the form               -> a new feature is created with
                                                  approval_status = "pending".
    2. A reviewer opens the feature layer in   -> changes approval_status to
       AGOL (or the form's "View answers")        "approved" or "rejected".
    3. The web app queries the layer with      -> only approved sites show up
       approval_status = 'approved'               on the map and in the cards.

Requirements
------------
    pip install "arcgis>=2.3" openpyxl

    A named user account on ArcGIS Online (or Enterprise) with permission to
    create content. Add-ons: Survey123 is included with every named user.

Usage
-----
    # Generate XLSForm + publish in one shot
    python scripts/create_survey.py --user yourname

    # Just write the XLSForm; publish later via Survey123 Connect / Web Designer
    python scripts/create_survey.py --xlsform-only --out troop380.xlsx

    # Use a stored AGOL profile instead of prompting for password
    #   (see: arcgis.gis.GIS profile parameter)
    python scripts/create_survey.py --profile my_agol_profile

Notes
-----
- The XLSForm is also saved to disk so you can edit it and re-publish later.
- `approval_status` is hidden from submitters via a calculate() expression so
  every new submission lands as "pending". Reviewers update the value in AGOL.
- If your `arcgis` install doesn't expose Survey.publish() (older versions),
  the script falls back to writing only the XLSForm and prints instructions
  for finishing the publish step in Survey123 Connect.
"""

from __future__ import annotations

import argparse
import getpass
import os
import sys
from typing import List, Tuple

# openpyxl is the only hard dep at script-generation time; arcgis is only
# imported when we actually publish (so --xlsform-only works without it).
from openpyxl import Workbook


# ─── Survey metadata ──────────────────────────────────────────────────────────

SURVEY_TITLE   = "Troop 380 — Family Camping Ideas"
SURVEY_SUMMARY = (
    "Share a camping or activity idea for Troop 380 / Pack 380. "
    "Submissions are reviewed before they appear on the Plan-Stuff-A-Thon "
    "planning guide."
)
SURVEY_TAGS    = ["Troop 380", "Scouting America", "BALOO", "Camping",
                  "Plan-Stuff-A-Thon"]
SURVEY_FORM_ID = "troop380_camping_ideas"
SURVEY_VERSION = "1.0.0"


# ─── XLSForm rows ─────────────────────────────────────────────────────────────
# Column order: type, name, label, required, default, calculation, hint, appearance

SurveyRow = Tuple[str, str, str, str, str, str, str, str]

SURVEY_ROWS: List[SurveyRow] = [
    # ── Submitter ────────────────────────────────────────────────────────────
    ("begin group", "submitter", "Your info",
        "", "", "", "", ""),
    ("text", "submitter_name",  "Your name",
        "yes", "", "", "First & last", ""),
    ("text", "submitter_email", "Your email",
        "yes", "", "", "Only used if we have questions", ""),
    ("text", "submitter_unit",  "Pack / Den / Patrol",
        "no",  "", "", "e.g. Pack 380, Wolf Den 2", ""),
    ("end group", "", "", "", "", "", "", ""),

    # ── Activity basics ──────────────────────────────────────────────────────
    ("begin group", "activity", "Activity idea",
        "", "", "", "", ""),
    ("text",     "title",       "Activity title",
        "yes", "", "", "Short, descriptive — e.g. 'Cunningham Falls family camp'", ""),
    ("text",     "description", "Short description",
        "yes", "", "", "1–3 sentences scouts and parents will read on the card", "multiline"),
    ("geopoint", "location",    "Location on the map",
        "yes", "", "", "Drop the pin where the activity happens", ""),

    ("select_one drive", "drive", "Drive time from DC",
        "yes", "near", "", "", "minimal"),
    ("select_one style", "style", "Trip style",
        "yes", "day",  "", "", "minimal"),

    ("select_multiple type",   "types", "Activity types",
        "yes", "",         "", "Pick all that apply", ""),
    ("select_multiple season", "seas",  "Seasons it works in",
        "yes", "Sp Su Fa", "", "Pick all that apply", ""),

    ("text", "cost",      "Approximate cost",
        "no", "", "", "e.g. ~$10/person, free, $30/site", ""),
    ("text", "cost_note", "Cost & logistics notes",
        "no", "", "", "Permits, group rates, parking, etc.", "multiline"),
    ("text", "website",   "Official website",
        "no", "", "", "https://...", ""),
    ("text", "reqs",      "Requirements / what to know",
        "no", "", "", "One per line works great", "multiline"),

    ("select_one yes_no", "baloo", "BALOO-compliant for Cub Scouts?",
        "yes", "no", "", "Suitable for Cub pack/den outings and recruiting events", ""),
    ("text",  "merits", "Merit badges or Cub adventures",
        "no", "", "", "Comma-separated", ""),
    ("image", "photo",  "Optional photo",
        "no", "", "", "Helps reviewers picture the site", ""),
    ("end group", "", "", "", "", "", "", ""),

    # ── Hidden: review state ─────────────────────────────────────────────────
    # Calculated to "pending" on every submission so submitters can't bypass
    # review. A reviewer updates the value to "approved" or "rejected" in AGOL.
    ("calculate", "approval_status", "",
        "", "", "'pending'", "", ""),
]


# ─── Choice lists ─────────────────────────────────────────────────────────────

CHOICES: List[Tuple[str, str, str]] = [
    ("drive", "near",   "Under 2 hrs from DC"),
    ("drive", "mid",    "2 – 3.5 hrs"),
    ("drive", "far",    "4 – 5 hrs"),
    ("drive", "longer", "Over 5 hrs"),

    ("style", "day",       "Day trip"),
    ("style", "overnight", "Overnight"),
    ("style", "either",    "Could be either"),

    ("type", "water",    "Water"),
    ("type", "hiking",   "Hiking"),
    ("type", "climbing", "Climbing"),
    ("type", "cave",     "Caving"),
    ("type", "bike",     "Biking"),
    ("type", "beach",    "Beach"),
    ("type", "history",  "History"),
    ("type", "multi",    "Multi-activity"),

    ("season", "Sp", "Spring"),
    ("season", "Su", "Summer"),
    ("season", "Fa", "Fall"),
    ("season", "Wi", "Winter"),

    ("yes_no", "yes", "Yes"),
    ("yes_no", "no",  "No / not sure"),

    # Reviewer updates these values directly in AGOL after submission.
    ("approval", "pending",  "Pending review"),
    ("approval", "approved", "Approved (visible)"),
    ("approval", "rejected", "Not a fit"),
]


SURVEY_HEADER  = ["type", "name", "label", "required", "default",
                  "calculation", "hint", "appearance"]
CHOICES_HEADER = ["list_name", "name", "label"]
SETTINGS_HEADER = ["form_title", "form_id", "version", "instance_name"]


# ─── XLSForm builder ──────────────────────────────────────────────────────────

def write_xlsform(path: str) -> None:
    wb = Workbook()

    survey_ws = wb.active
    survey_ws.title = "survey"
    survey_ws.append(SURVEY_HEADER)
    for row in SURVEY_ROWS:
        survey_ws.append(list(row))

    choices_ws = wb.create_sheet("choices")
    choices_ws.append(CHOICES_HEADER)
    for row in CHOICES:
        choices_ws.append(list(row))

    settings_ws = wb.create_sheet("settings")
    settings_ws.append(SETTINGS_HEADER)
    settings_ws.append([SURVEY_TITLE, SURVEY_FORM_ID, SURVEY_VERSION,
                        # instance_name = readable record label in feature layer
                        "concat(${title}, ' — ', ${submitter_name})"])

    wb.save(path)


# ─── Publish to ArcGIS Online ─────────────────────────────────────────────────

def publish_to_agol(gis, xlsx_path: str):
    """Create the Survey123 form item and publish from the XLSForm.

    Returns the Survey object on success. Falls back gracefully (returns
    None) if the installed `arcgis` version doesn't expose Survey.publish().
    """
    from arcgis.apps.survey123 import SurveyManager

    sm = SurveyManager(gis)
    survey = sm.create(
        title=SURVEY_TITLE,
        summary=SURVEY_SUMMARY,
        tags=SURVEY_TAGS,
    )

    publish_fn = getattr(survey, "publish", None)
    if publish_fn is None:
        print("  ! Survey.publish() is not available in your `arcgis` version.")
        print("    The empty form item was created; finish publishing by")
        print(f"    dropping {xlsx_path} into Survey123 Connect.")
        return survey

    # Different arcgis versions name the parameter differently. Try the
    # common spellings.
    for kwarg in ("xls_path", "xlsx_path", "xlsform", "file"):
        try:
            publish_fn(**{kwarg: xlsx_path})
            return survey
        except TypeError:
            continue

    print("  ! Could not figure out the publish() parameter name on this version.")
    print(f"    Open {xlsx_path} in Survey123 Connect and click Publish.")
    return survey


# ─── CLI entry point ──────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(
        description="Create the Troop 380 family camping-ideas Survey123 form.")
    p.add_argument("--user", help="ArcGIS Online username (omit if --profile is set)")
    p.add_argument("--profile", help="Stored ArcGIS API for Python profile name")
    p.add_argument("--portal", default="https://www.arcgis.com",
                   help="Portal URL (default: ArcGIS Online)")
    p.add_argument("--out", default="troop380_camping_ideas.xlsx",
                   help="Where to write the XLSForm (default: %(default)s)")
    p.add_argument("--xlsform-only", action="store_true",
                   help="Only write the XLSForm; skip the AGOL publish step.")
    args = p.parse_args()

    print(f"Writing XLSForm  →  {args.out}")
    write_xlsform(args.out)

    if args.xlsform_only:
        print("Done (--xlsform-only). Drop the file into Survey123 Connect or")
        print("upload via Survey123 Web Designer to finish publishing.")
        return 0

    if not args.profile and not args.user:
        p.error("either --user or --profile is required (or pass --xlsform-only)")

    print("Connecting to ArcGIS Online…")
    from arcgis.gis import GIS  # imported lazily so --xlsform-only doesn't need it
    if args.profile:
        gis = GIS(profile=args.profile)
    else:
        pwd = os.environ.get("AGOL_PASSWORD") \
            or getpass.getpass(f"Password for {args.user}: ")
        gis = GIS(args.portal, args.user, pwd)
    print(f"  Signed in as {gis.users.me.username} on {gis.url}")

    print("Publishing survey + response feature service…")
    survey = publish_to_agol(gis, args.out)

    print()
    print("Done.")
    if survey is not None:
        item_id = (getattr(survey, "id", None)
                   or getattr(getattr(survey, "properties", {}), "get", lambda _: None)("id"))
        if item_id:
            base = gis.url.rstrip("/")
            print(f"  Form item       {base}/home/item.html?id={item_id}")
            print(f"  Form designer   https://survey123.arcgis.com/surveys/{item_id}/design")
            print(f"  Submit URL      https://survey123.arcgis.com/share/{item_id}")
            print(f"  View answers    https://survey123.arcgis.com/surveys/{item_id}/analyze")

    print()
    print("Next steps")
    print("  1. Open the form in Survey123 Web Designer and tweak look/feel.")
    print("  2. Share the form (publicly or with your scouting org).")
    print("  3. As submissions arrive, set approval_status = 'approved' on the")
    print("     ones to publish (via the layer's attribute table in AGOL).")
    print("  4. In data.js, swap the static ACTS array for a fetch from the")
    print("     response feature service with a where-clause:")
    print("        approval_status='approved'")
    print("     and map fields back to the existing card schema.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
